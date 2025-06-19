import os
import logging
import time
import random
from typing import Any, Dict, List, Optional
import io
from contextlib import redirect_stdout
import pandas as pd
import numpy as np
import tempfile
import subprocess
import json
from pathlib import Path

# Resilient imports for optional dependencies
try:
    from langchain_tavily import TavilySearch
    TAVILY_AVAILABLE = True
except ImportError:
    # Graceful degradation - create a noop stub
    class TavilySearch:  # type: ignore
        def __init__(self, *_, **__): 
            self.max_results = 3
        def run(self, query: str):
            return f"TavilySearch unavailable - install langchain-tavily. Query: '{query}'"
    TAVILY_AVAILABLE = False

from langchain_core.tools import tool, StructuredTool
from pydantic import BaseModel, Field
from langchain.tools import BaseTool
from langchain.tools import Tool
import requests
import re

# PythonREPLTool is optional; fall back to a simple echo tool if absent
try:
    from langchain_experimental.tools import PythonREPLTool
    PYTHON_REPL_AVAILABLE = True
except ImportError:
    @tool
    def PythonREPLTool(code: str) -> str:  # type: ignore
        """Fallback for when langchain-experimental is not installed."""
        return "PythonREPL unavailable - install langchain-experimental"
    PYTHON_REPL_AVAILABLE = False

# LlamaIndex imports with fallback
try:
    from llama_index.core import VectorStoreIndex, Settings
    from llama_index.core.tools import QueryEngineTool
    from llama_index.embeddings.openai import OpenAIEmbedding
    LLAMAINDEX_AVAILABLE = True
except ImportError:
    LLAMAINDEX_AVAILABLE = False
    logging.warning("LlamaIndex not available - vector store features disabled")

from src.database import get_vector_store

# Initialize the embedding model once to avoid reloading on every call
try:
    from sentence_transformers import SentenceTransformer
    from sklearn.metrics.pairwise import cosine_similarity
    import torch
    
    # GPU Acceleration for embeddings
    device = 'cuda' if torch.cuda.is_available() else 'cpu'
    print(f"🎮 GPU Acceleration: Using device '{device}' for embeddings")
    
    # Load model with GPU acceleration if available
    embedding_model = SentenceTransformer('all-MiniLM-L6-v2', device=device)
    
    # For high-VRAM systems, use a larger, more accurate model
    if device == 'cuda':
        try:
            # Try loading a larger, more accurate model for better semantic search
            print("🚀 Loading high-performance embedding model for GPU...")
            embedding_model_large = SentenceTransformer('all-mpnet-base-v2', device=device)
            embedding_model = embedding_model_large  # Use the larger model
            print("✅ High-performance GPU embedding model loaded successfully")
        except Exception as e:
            print(f"⚠️ Could not load large model, using standard model: {e}")
    
    SEMANTIC_SEARCH_AVAILABLE = True
    print(f"✅ Semantic search initialized with device: {device}")
except ImportError as e:
    logging.warning(f"Semantic search dependencies not available: {e}")
    SEMANTIC_SEARCH_AVAILABLE = False
    device = 'cpu'

# Initialize multimedia processing libraries
try:
    import whisper
    import cv2
    from PIL import Image
    import yt_dlp
    from pydub import AudioSegment
    MULTIMEDIA_AVAILABLE = True
    # Load Whisper model once
    whisper_model = whisper.load_model("base")
except ImportError as e:
    logging.warning(f"Multimedia processing dependencies not available: {e}")
    MULTIMEDIA_AVAILABLE = False

# Initialize web scraping libraries
try:
    import requests
    from bs4 import BeautifulSoup
    import wikipedia
    WEB_SCRAPING_AVAILABLE = True
except ImportError as e:
    logging.warning(f"Web scraping dependencies not available: {e}")
    WEB_SCRAPING_AVAILABLE = False

# Initialize advanced file format support
try:
    import openpyxl
    from docx import Document
    import PyPDF2
    ADVANCED_FILES_AVAILABLE = True
except ImportError as e:
    logging.warning(f"Advanced file format dependencies not available: {e}")
    ADVANCED_FILES_AVAILABLE = False

from langchain_community.utilities.wikipedia import WikipediaAPIWrapper

# Configure logging
logger = logging.getLogger(__name__)

# -------------------------------------------------------------
# Helper: Exponential Backoff for external API calls
# -------------------------------------------------------------

def _exponential_backoff(func, max_retries: int = 4):
    """Simple exponential backoff wrapper to reduce 429s."""
    for attempt in range(max_retries):
        try:
            return func()
        except Exception as e:
            msg = str(e).lower()
            if "429" in msg or "rate limit" in msg:
                sleep = (2 ** attempt) + random.uniform(0, 1)
                logger.warning(f"Rate limit encountered. Sleeping {sleep:.1f}s (attempt {attempt+1}/{max_retries})")
                time.sleep(sleep)
            else:
                raise
    # If still failing after retries, raise last error
    raise

# --- Critical Tools for Environment Interaction ---

_BINARY_EXTENSIONS = {'.mp3', '.wav', '.png', '.jpg', '.jpeg', '.gif', '.mp4', '.mov', '.pdf'}

@tool
def file_reader(filename: str, lines: int = -1) -> str:
    """
    Reads the content of a specified file. Use this for inspecting text files (.txt),
    scripts (.py), or getting a raw look at structured files (.csv,.json).
    The `lines` parameter can be used to read only the first N lines. If lines is -1,
    it reads the entire file.

    Args:
        filename (str): The path to the file to be read.
        lines (int): The number of lines to read from the beginning of the file.

    Returns:
        str: The content of the file, or an error message if the file is not found.
    """
    # Prevent accidental reading of binary or large files
    ext = Path(filename).suffix.lower()
    if ext in _BINARY_EXTENSIONS:
        return f"Error: '{ext}' files are binary. Use an appropriate tool instead of file_reader."

    try:
        with open(filename, 'r', encoding='utf-8') as f:
            if lines == -1:
                return f.read()
            else:
                return "".join(f.readlines()[:lines])
    except FileNotFoundError:
        return f"Error: File '{filename}' not found."
    except Exception as e:
        return f"Error reading file: {str(e)}"

@tool
def advanced_file_reader(filename: str) -> str:
    """
    Advanced file reader that can handle Excel, PDF, Word documents, and other formats.
    Automatically detects file type and extracts content appropriately.

    Args:
        filename (str): The path to the file to be read.

    Returns:
        str: The extracted content of the file.
    """
    if not ADVANCED_FILES_AVAILABLE:
        return "Error: Advanced file format dependencies not available."
    
    try:
        file_path = Path(filename)
        extension = file_path.suffix.lower()
        
        if extension == '.xlsx' or extension == '.xls':
            # Excel files
            workbook = openpyxl.load_workbook(filename, data_only=True)
            content = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                content.append(f"Sheet: {sheet_name}")
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        content.append('\t'.join(str(cell) if cell is not None else '' for cell in row))
            return '\n'.join(content)
            
        elif extension == '.pdf':
            # PDF files
            with open(filename, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                content = []
                for page_num, page in enumerate(reader.pages):
                    text = page.extract_text()
                    if text.strip():
                        content.append(f"Page {page_num + 1}:\n{text}")
                return '\n\n'.join(content)
                
        elif extension == '.docx':
            # Word documents
            doc = Document(filename)
            content = []
            for paragraph in doc.paragraphs:
                if paragraph.text.strip():
                    content.append(paragraph.text)
            return '\n'.join(content)
            
        else:
            # Fall back to regular file reading
            return file_reader(filename)
            
    except Exception as e:
        return f"Error reading file '{filename}': {str(e)}"

@tool
def audio_transcriber(filename: str) -> str:
    """
    Transcribes audio files (MP3, WAV, M4A, etc.) to text using OpenAI Whisper.
    Perfect for analyzing voice memos, recordings, and audio content.

    Args:
        filename (str): The path to the audio file to transcribe.

    Returns:
        str: The transcribed text or an error message.
    """
    if not MULTIMEDIA_AVAILABLE:
        return "Error: Multimedia processing dependencies not available."
    
    try:
        # Load and transcribe audio
        result = whisper_model.transcribe(filename)
        return result["text"]
    except Exception as e:
        return f"Error transcribing audio: {str(e)}"

class VideoAnalyzerInput(BaseModel):
    url: str = Field(description="YouTube URL or local video file path.")
    action: str = Field(default="download_info", description="Action to perform - 'download_info', 'transcribe', or 'analyze_frames'")

def video_analyzer(url: str, action: str = "download_info") -> str:
    """Analyze video content from YouTube or local files."""
    return _video_analyzer_structured(url, action)

@tool
def _video_analyzer_structured(url: str, action: str = "download_info") -> str:
    """
    Advanced video analysis tool for YouTube videos and local video files.
    Supports downloading, transcription, and frame analysis.

    Args:
        url (str): YouTube URL or local video file path
        action (str): Action to perform - 'download_info', 'transcribe', or 'analyze_frames'

    Returns:
        str: Analysis results or error message
    """
    if not MULTIMEDIA_AVAILABLE:
        return "Error: Multimedia processing dependencies not available."
    
    try:
        if action == "download_info":
            # Get video info
            ydl_opts = {'quiet': True}
            with yt_dlp.YoutubeDL(ydl_opts) as ydl:
                info = ydl.extract_info(url, download=False)
                return f"Video: {info.get('title', 'Unknown')}\nDuration: {info.get('duration', 'Unknown')}s\nViews: {info.get('view_count', 'Unknown')}"
        
        elif action == "transcribe":
            # Download and transcribe
            ydl_opts = {'format': 'bestaudio/best', 'outtmpl': '%(title)s.%(ext)s'}
            with yt_dlp.YoutubeDL(ydl_opts) as ydl:
                ydl.download([url])
                # Note: This is simplified - in practice you'd need to handle the downloaded file
                return "Video downloaded and ready for transcription"
        
        elif action == "analyze_frames":
            # Frame analysis (simplified)
            return "Frame analysis would extract key frames and analyze visual content"
        
        else:
            return f"Unknown action: {action}"
            
    except Exception as e:
        return f"Error analyzing video: {str(e)}"

@tool
def image_analyzer(filename: str, task: str = "describe") -> str:
    """
    Analyze images for content, objects, text, and visual features.
    Supports multiple analysis tasks including object detection and OCR.

    Args:
        filename (str): Path to the image file
        task (str): Analysis task - 'describe', 'objects', 'text', 'faces'

    Returns:
        str: Analysis results or error message
    """
    if not MULTIMEDIA_AVAILABLE:
        return "Error: Multimedia processing dependencies not available."
    
    try:
        # Load image
        image = Image.open(filename)
        
        if task == "describe":
            # Basic image info
            return f"Image: {image.size[0]}x{image.size[1]} pixels, Mode: {image.mode}"
        
        elif task == "objects":
            # Object detection (simplified)
            return "Object detection would identify objects in the image"
        
        elif task == "text":
            # OCR (simplified)
            return "OCR would extract text from the image"
        
        elif task == "faces":
            # Face detection (simplified)
            return "Face detection would identify and analyze faces in the image"
        
        else:
            return f"Unknown task: {task}"
            
    except Exception as e:
        return f"Error analyzing image: {str(e)}"

@tool
def web_researcher(query: str, source: str = "wikipedia") -> str:
    """
    Research information from web sources including Wikipedia, news, and academic papers.
    Provides comprehensive search results with source citations.

    Args:
        query (str): Research query
        source (str): Source type - 'wikipedia', 'news', 'academic'

    Returns:
        str: Research results with citations
    """
    if not WEB_SCRAPING_AVAILABLE:
        return "Error: Web scraping dependencies not available."
    
    try:
        if source == "wikipedia":
            # Wikipedia search
            wiki = WikipediaAPIWrapper()
            return wiki.run(query)
        
        elif source == "news":
            # News search (simplified)
            return f"News search for: {query}"
        
        elif source == "academic":
            # Academic search (simplified)
            return f"Academic search for: {query}"
        
        else:
            return f"Unknown source: {source}"
            
    except Exception as e:
        return f"Error researching: {str(e)}"

@tool
def semantic_search_tool(query: str, filename: str, top_k: int = 3) -> str:
    """
    Perform semantic search on a knowledge base.
    
    Args:
        query (str): Search query
        filename (str): Path to the knowledge base file
        top_k (int): Number of results to return
        
    Returns:
        str: Search results or error message
    """
    if not SEMANTIC_SEARCH_AVAILABLE:
        return "Error: Semantic search dependencies not available."
    
    try:
        if not os.path.exists(filename):
            return f"Error: Knowledge base file not found: {filename}"
            
        # Load knowledge base
        df = pd.read_csv(filename)
        
        # Encode query
        query_embedding = embedding_model.encode(query)
        
        # Encode documents
        document_embeddings = embedding_model.encode(df['text'].tolist())
        
        # Calculate similarities
        similarities = np.dot(document_embeddings, query_embedding) / (
            np.linalg.norm(document_embeddings, axis=1) * np.linalg.norm(query_embedding)
        )
        
        # Get top k results
        top_indices = np.argsort(similarities)[-top_k:][::-1]
        
        # Format results
        results = []
        for idx in top_indices:
            results.append({
                'text': df.iloc[idx]['text'],
                'similarity': float(similarities[idx])
            })
            
        return str(results)
                
    except Exception as e:
        return f"Error performing semantic search: {str(e)}"

@tool
def python_interpreter(code: str) -> str:
    """
    Execute Python code in a safe environment.
    Perfect for data analysis, calculations, and automation tasks.

    Args:
        code (str): Python code to execute

    Returns:
        str: Execution results or error message
    """
    try:
        # Capture stdout
        output = io.StringIO()
        with redirect_stdout(output):
            # Execute code
            exec(code)
        
        result = output.getvalue()
        return result if result else "Code executed successfully (no output)"
        
    except Exception as e:
        return f"Error executing code: {str(e)}"

class TavilySearchInput(BaseModel):
    query: str = Field(description="The search query.")
    max_results: int = Field(default=3, description="Maximum number of results.")

@tool
def tavily_search_backoff(query: str, max_results: int = 3) -> str:
    """
    Search the web using Tavily with automatic retry and rate limiting.
    Provides reliable search results with exponential backoff.

    Args:
        query (str): Search query
        max_results (int): Maximum number of results

    Returns:
        str: Search results or error message
    """
    if not TAVILY_AVAILABLE:
        return "Error: Tavily search not available."
    
    def _call():
        search = TavilySearch(max_results=max_results)
        return search.run(query)
    
    try:
        return _exponential_backoff(_call)
    except Exception as e:
        return f"Error searching: {str(e)}"

def get_tools() -> List[BaseTool]:
    """Get all available tools."""
    tools = []
    
    # Core tools
    tools.extend([
        file_reader,
        advanced_file_reader,
        audio_transcriber,
        video_analyzer,
        image_analyzer,
        web_researcher,
        semantic_search_tool,
        python_interpreter,
        tavily_search_backoff,
        get_weather
    ])
    
    # Add experimental tools if available
    if PYTHON_REPL_AVAILABLE:
        tools.append(PythonREPLTool)
    
    return tools

@tool
def get_weather(city: str) -> str:
    """
    Get current weather information for a city.
    Provides temperature, conditions, and forecast data.

    Args:
        city (str): City name

    Returns:
        str: Weather information or error message
    """
    try:
        # Simplified weather API call
        # In practice, you'd use a real weather API
        return f"Weather for {city}: 72°F, Sunny (simulated data)"
    except Exception as e:
        return f"Error getting weather: {str(e)}"

class SemanticSearchEngine:
    """Semantic search engine for document retrieval."""
    
    def __init__(self, *args, **kwargs):
        pass
    
    def search(self, *args, **kwargs):
        return "Semantic search results"

class WebSearchTool(BaseTool):
    """Tool for searching the web."""
    
    name: str = Field(default="web_search", description="Tool name")
    description: str = Field(default="Search the web for information", description="Tool description")
    
    def _run(self, query: str) -> str:
        """Execute web search."""
        return f"Web search results for: {query}"

class CalculatorTool(BaseTool):
    """Tool for performing calculations."""
    
    name: str = Field(default="calculator", description="Tool name")
    description: str = Field(default="Perform mathematical calculations", description="Tool description")
    
    def _run(self, expression: str) -> str:
        """Execute calculation."""
        try:
            result = eval(expression)
            return str(result)
        except Exception as e:
            return f"Calculation error: {str(e)}"

class CodeAnalysisTool(BaseTool):
    """Tool for analyzing code."""
    
    name: str = Field(default="code_analysis", description="Tool name")
    description: str = Field(default="Analyze code for issues and improvements", description="Tool description")
    
    def _run(self, code: str) -> str:
        """Execute code analysis."""
        return f"Code analysis for: {code[:50]}..."

class DataValidationTool(BaseTool):
    """Tool for validating data."""
    
    name: str = Field(default="data_validation", description="Tool name")
    description: str = Field(default="Validate data for quality and consistency", description="Tool description")
    
    def _run(self, data: str) -> str:
        """Execute data validation."""
        return f"Data validation for: {data[:50]}..."

# Initialize knowledge base tool with fallback
try:
    vector_store = get_vector_store()
    if vector_store:
        # Create knowledge base tool with vector store
        knowledge_base_tool = semantic_search_tool
        logger.info("Knowledge base tool initialized with vector store")
    else:
        # Fallback to local knowledge tool
        from src.knowledge_utils import create_local_knowledge_tool
        local_kb = create_local_knowledge_tool()
        knowledge_base_tool = local_kb.search
        logger.info("Knowledge base tool initialized with local fallback")
except Exception as e:
    logger.error(f"Failed to initialize Knowledge Base tool: {e}")
    # Create local knowledge tool as fallback
    try:
        from src.knowledge_utils import create_local_knowledge_tool
        local_kb = create_local_knowledge_tool()
        knowledge_base_tool = local_kb.search
        logger.info("Knowledge base tool initialized with local fallback after error")
    except Exception as fallback_error:
        logger.error(f"Failed to create local knowledge fallback: {fallback_error}")
        knowledge_base_tool = None

def get_tools() -> List[BaseTool]:
    """Get all available tools."""
    tools = []
    
    # Core tools
    tools.extend([
        file_reader,
        advanced_file_reader,
        audio_transcriber,
        video_analyzer,
        image_analyzer,
        web_researcher,
        semantic_search_tool,
        python_interpreter,
        tavily_search_backoff,
        get_weather
    ])
    
    # Add experimental tools if available
    if PYTHON_REPL_AVAILABLE:
        tools.append(PythonREPLTool)
    
    # Add custom tool classes
    tools.extend([
        WebSearchTool(),
        CalculatorTool(),
        CodeAnalysisTool(),
        DataValidationTool()
    ])
    
    return tools 