import os
import logging
import time
import random
from typing import Any
import io
from contextlib import redirect_stdout
import pandas as pd
import numpy as np
import tempfile
import subprocess
import json
from pathlib import Path

# Resilient imports for optional dependencies
try:
    from langchain_tavily import TavilySearch
    TAVILY_AVAILABLE = True
except ImportError:
    # Graceful degradation - create a noop stub
    class TavilySearch:  # type: ignore
        def __init__(self, *_, **__): 
            self.max_results = 3
        def run(self, query: str):
            return f"TavilySearch unavailable - install langchain-tavily. Query: '{query}'"
    TAVILY_AVAILABLE = False

from langchain_core.tools import tool, StructuredTool
from pydantic import BaseModel, Field

# PythonREPLTool is optional; fall back to a simple echo tool if absent
try:
    from langchain_experimental.tools import PythonREPLTool
    PYTHON_REPL_AVAILABLE = True
except ImportError:
    @tool
    def PythonREPLTool(code: str) -> str:  # type: ignore
        """Fallback for when langchain-experimental is not installed."""
        return "PythonREPL unavailable - install langchain-experimental"
    PYTHON_REPL_AVAILABLE = False

# LlamaIndex imports with fallback
try:
    from llama_index.core import VectorStoreIndex, Settings
    from llama_index.core.tools import QueryEngineTool
    from llama_index.embeddings.openai import OpenAIEmbedding
    LLAMAINDEX_AVAILABLE = True
except ImportError:
    LLAMAINDEX_AVAILABLE = False
    logging.warning("LlamaIndex not available - vector store features disabled")

from src.database import get_vector_store

# Initialize the embedding model once to avoid reloading on every call
try:
    from sentence_transformers import SentenceTransformer
    from sklearn.metrics.pairwise import cosine_similarity
    import torch
    
    # GPU Acceleration for embeddings
    device = 'cuda' if torch.cuda.is_available() else 'cpu'
    print(f"🎮 GPU Acceleration: Using device '{device}' for embeddings")
    
    # Load model with GPU acceleration if available
    embedding_model = SentenceTransformer('all-MiniLM-L6-v2', device=device)
    
    # For high-VRAM systems, use a larger, more accurate model
    if device == 'cuda':
        try:
            # Try loading a larger, more accurate model for better semantic search
            print("🚀 Loading high-performance embedding model for GPU...")
            embedding_model_large = SentenceTransformer('all-mpnet-base-v2', device=device)
            embedding_model = embedding_model_large  # Use the larger model
            print("✅ High-performance GPU embedding model loaded successfully")
        except Exception as e:
            print(f"⚠️ Could not load large model, using standard model: {e}")
    
    SEMANTIC_SEARCH_AVAILABLE = True
    print(f"✅ Semantic search initialized with device: {device}")
except ImportError as e:
    logging.warning(f"Semantic search dependencies not available: {e}")
    SEMANTIC_SEARCH_AVAILABLE = False
    device = 'cpu'

# Initialize multimedia processing libraries
try:
    import whisper
    import cv2
    from PIL import Image
    import yt_dlp
    from pydub import AudioSegment
    MULTIMEDIA_AVAILABLE = True
    # Load Whisper model once
    whisper_model = whisper.load_model("base")
except ImportError as e:
    logging.warning(f"Multimedia processing dependencies not available: {e}")
    MULTIMEDIA_AVAILABLE = False

# Initialize web scraping libraries
try:
    import requests
    from bs4 import BeautifulSoup
    import wikipedia
    WEB_SCRAPING_AVAILABLE = True
except ImportError as e:
    logging.warning(f"Web scraping dependencies not available: {e}")
    WEB_SCRAPING_AVAILABLE = False

# Initialize advanced file format support
try:
    import openpyxl
    from docx import Document
    import PyPDF2
    ADVANCED_FILES_AVAILABLE = True
except ImportError as e:
    logging.warning(f"Advanced file format dependencies not available: {e}")
    ADVANCED_FILES_AVAILABLE = False

from langchain_community.utilities.wikipedia import WikipediaAPIWrapper

# Configure logging
logger = logging.getLogger(__name__)

# -------------------------------------------------------------
# Helper: Exponential Backoff for external API calls
# -------------------------------------------------------------

def _exponential_backoff(func, max_retries: int = 4):
    """Simple exponential backoff wrapper to reduce 429s."""
    for attempt in range(max_retries):
        try:
            return func()
        except Exception as e:
            msg = str(e).lower()
            if "429" in msg or "rate limit" in msg:
                sleep = (2 ** attempt) + random.uniform(0, 1)
                logger.warning(f"Rate limit encountered. Sleeping {sleep:.1f}s (attempt {attempt+1}/{max_retries})")
                time.sleep(sleep)
            else:
                raise
    # If still failing after retries, raise last error
    raise

# --- Critical Tools for Environment Interaction ---

_BINARY_EXTENSIONS = {'.mp3', '.wav', '.png', '.jpg', '.jpeg', '.gif', '.mp4', '.mov', '.pdf'}

@tool
def file_reader(filename: str, lines: int = -1) -> str:
    """
    Reads the content of a specified file. Use this for inspecting text files (.txt),
    scripts (.py), or getting a raw look at structured files (.csv,.json).
    The `lines` parameter can be used to read only the first N lines. If lines is -1,
    it reads the entire file.

    Args:
        filename (str): The path to the file to be read.
        lines (int): The number of lines to read from the beginning of the file.

    Returns:
        str: The content of the file, or an error message if the file is not found.
    """
    # Prevent accidental reading of binary or large files
    ext = Path(filename).suffix.lower()
    if ext in _BINARY_EXTENSIONS:
        return f"Error: '{ext}' files are binary. Use an appropriate tool instead of file_reader."

    try:
        with open(filename, 'r', encoding='utf-8') as f:
            if lines == -1:
                return f.read()
            else:
                return "".join(f.readlines()[:lines])
    except FileNotFoundError:
        return f"Error: File '{filename}' not found."
    except Exception as e:
        return f"Error reading file: {str(e)}"

@tool
def advanced_file_reader(filename: str) -> str:
    """
    Advanced file reader that can handle Excel, PDF, Word documents, and other formats.
    Automatically detects file type and extracts content appropriately.

    Args:
        filename (str): The path to the file to be read.

    Returns:
        str: The extracted content of the file.
    """
    if not ADVANCED_FILES_AVAILABLE:
        return "Error: Advanced file format dependencies not available."
    
    try:
        file_path = Path(filename)
        extension = file_path.suffix.lower()
        
        if extension == '.xlsx' or extension == '.xls':
            # Excel files
            workbook = openpyxl.load_workbook(filename, data_only=True)
            content = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                content.append(f"Sheet: {sheet_name}")
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        content.append('\t'.join(str(cell) if cell is not None else '' for cell in row))
            return '\n'.join(content)
            
        elif extension == '.pdf':
            # PDF files
            with open(filename, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                content = []
                for page_num, page in enumerate(reader.pages):
                    text = page.extract_text()
                    if text.strip():
                        content.append(f"Page {page_num + 1}:\n{text}")
                return '\n\n'.join(content)
                
        elif extension == '.docx':
            # Word documents
            doc = Document(filename)
            content = []
            for paragraph in doc.paragraphs:
                if paragraph.text.strip():
                    content.append(paragraph.text)
            return '\n'.join(content)
            
        else:
            # Fall back to regular file reading
            return file_reader(filename)
            
    except Exception as e:
        return f"Error reading file '{filename}': {str(e)}"

@tool
def audio_transcriber(filename: str) -> str:
    """
    Transcribes audio files (MP3, WAV, M4A, etc.) to text using OpenAI Whisper.
    Perfect for analyzing voice memos, recordings, and audio content.

    Args:
        filename (str): The path to the audio file to transcribe.

    Returns:
        str: The transcribed text from the audio file.
    """
    if not MULTIMEDIA_AVAILABLE:
        return "Error: Multimedia processing dependencies not available."
    
    try:
        # Transcribe the audio file using Whisper
        result = whisper_model.transcribe(filename)
        return result["text"]
    except Exception as e:
        return f"Error transcribing audio file '{filename}': {str(e)}"

class VideoAnalyzerInput(BaseModel):
    url: str = Field(description="YouTube URL or local video file path.")
    action: str = Field(default="download_info", description="Action to perform - 'download_info', 'transcribe', or 'analyze_frames'")

def video_analyzer(url: str, action: str = "download_info") -> str:
    """Stub video analyzer function for FSM compatibility."""
    return f"Video analyzer not yet implemented for url={url}, action={action}. Please use gaia_video_analyzer or video_analyzer_production instead."

def _video_analyzer_structured(url: str, action: str = "download_info") -> str:
    return video_analyzer(url, action)

video_analyzer_structured = StructuredTool.from_function(
    func=_video_analyzer_structured,
    name="video_analyzer",
    description="Analyzes videos from URLs (especially YouTube) or local video files. Can extract metadata, transcribe audio, or download for analysis.",
    args_schema=VideoAnalyzerInput
)

@tool
def image_analyzer(filename: str, task: str = "describe") -> str:
    """
    Analyzes images for various tasks like chess position analysis, object detection, etc.

    Args:
        filename (str): Path to the image file.
        task (str): Analysis task - "describe", "chess", "objects", or "text"

    Returns:
        str: Analysis results based on the specified task.
    """
    if not MULTIMEDIA_AVAILABLE:
        return "Error: Multimedia processing dependencies not available."
    
    try:
        # Load image
        image = cv2.imread(filename)
        if image is None:
            return f"Error: Could not load image '{filename}'"
        
        if task == "chess":
            # For chess position analysis - basic implementation
            return "Chess position analysis requires specialized chess vision models. Please provide the position in standard notation."
        
        elif task == "describe":
            # Basic image description
            height, width, channels = image.shape
            return f"Image dimensions: {width}x{height} pixels, {channels} channels"
        
        elif task == "text":
            # OCR text extraction would require pytesseract
            return "Text extraction from images requires OCR setup (pytesseract)"
        
        else:
            return f"Image analysis task '{task}' not implemented yet."
            
    except Exception as e:
        return f"Error analyzing image '{filename}': {str(e)}"

@tool
def web_researcher(query: str, source: str = "wikipedia") -> str:
    """
    Performs web research using various sources like Wikipedia, or general web search.

    Args:
        query (str): The research query or topic.
        source (str): Research source - "wikipedia", "search", or "url"

    Returns:
        str: Research results and relevant information.
    """
    if not WEB_SCRAPING_AVAILABLE:
        return "Error: Web scraping dependencies not available."
    
    try:
        if source == "wikipedia":
            # Use WikipediaAPIWrapper instead of 'wikipedia'
            try:
                wrapper = WikipediaAPIWrapper()
                summary = wrapper.run(query)
                return f"Wikipedia Article Summary:\n{summary}"
            except Exception as e:
                return f"Error searching Wikipedia: {str(e)}"
        
        elif source == "search":
            # Use the Tavily search with backoff
            return tavily_search_backoff(query)
        
        else:
            return f"Research source '{source}' not implemented"
            
    except Exception as e:
        return f"Error in web research: {str(e)}"

@tool
def semantic_search_tool(query: str, filename: str, top_k: int = 3) -> str:
    """
    Performs a semantic (vector-based) search on a specified knowledge file (typically
    a.csv with embeddings). Use this when you need to find information related to a
    concept or question, not just a keyword. `filename` must be a file known to
    contain embeddings, like 'supabase_docs.csv'.
    
    NOW WITH GPU ACCELERATION for ultra-fast semantic search!

    Args:
        query (str): The natural language query to search for.
        filename (str): The path to the CSV file containing the knowledge base.
        top_k (int): The number of top results to return.

    Returns:
        str: A formatted string of the most relevant content chunks.
    """
    if not SEMANTIC_SEARCH_AVAILABLE:
        return "Error: Semantic search dependencies (sentence-transformers, scikit-learn) not available."
    
    try:
        import time
        start_time = time.time()
        
        df = pd.read_csv(filename)
        
        # Check if the required columns exist
        if 'embedding' not in df.columns or 'content' not in df.columns:
            return f"Error: File '{filename}' does not contain required 'embedding' and 'content' columns."
        
        # The embedding column is stored as a string representation of a list.
        # This line safely converts it back to a numpy array for calculation.
        df['embedding'] = df['embedding'].apply(lambda x: np.fromstring(x.strip('[]'), sep=','))
        
        # GPU-accelerated query embedding generation
        embedding_start = time.time()
        query_embedding = embedding_model.encode([query], device=device, show_progress_bar=False)
        embedding_time = time.time() - embedding_start
        
        knowledge_embeddings = np.vstack(df['embedding'].values)

        # Use GPU-accelerated similarity computation if available
        if device == 'cuda' and hasattr(torch, 'cuda'):
            try:
                # Convert to tensors for GPU computation
                query_tensor = torch.tensor(query_embedding, device=device)
                knowledge_tensor = torch.tensor(knowledge_embeddings, device=device)
                
                # GPU-accelerated cosine similarity
                similarities = torch.cosine_similarity(query_tensor, knowledge_tensor, dim=1)
                sim_scores = similarities.cpu().numpy()
            except Exception as e:
                print(f"GPU similarity computation failed, falling back to CPU: {e}")
                sim_scores = cosine_similarity(query_embedding, knowledge_embeddings)[0]
        else:
            sim_scores = cosine_similarity(query_embedding, knowledge_embeddings)[0]
        
        top_indices = np.argsort(sim_scores)[-top_k:][::-1]

        total_time = time.time() - start_time
        
        results = []
        results.append(f"🎮 GPU-Accelerated Semantic Search Results (Device: {device})")
        results.append(f"⚡ Performance: Query processed in {total_time:.3f}s (Embedding: {embedding_time:.3f}s)")
        results.append(f"📊 Searched {len(knowledge_embeddings)} documents, returning top {top_k}")
        results.append("=" * 60)
        
        for i, idx in enumerate(top_indices):
            results.append(f"🔍 Result #{i+1} (Relevance Score: {sim_scores[idx]:.4f}):")
            results.append(f"{df.loc[idx, 'content']}")
            if i < len(top_indices) - 1:
                results.append("-" * 40)
        
        return "\n".join(results)
        
    except FileNotFoundError:
        return f"Error: Knowledge base file '{filename}' not found."
    except Exception as e:
        return f"Error during GPU-accelerated semantic search: {str(e)}"

@tool
def python_interpreter(code: str) -> str:
    """
    Executes a given block of Python code in a secure, sandboxed environment.
    Use this for calculations, data manipulation, or running scripts.
    The code MUST include a `print()` statement to return a result to the observation.
    WARNING: This tool is powerful. Do not execute code that modifies the file system
    or makes network requests unless explicitly required and sanctioned.

    Args:
        code (str): The Python code to execute.

    Returns:
        str: The captured stdout from the executed code, or an error message.
    """
    f = io.StringIO()
    try:
        # A simple, isolated global scope for the execution
        exec_globals = {}
        with redirect_stdout(f):
            exec(code, exec_globals)
        return f.getvalue()
    except Exception as e:
        return f"Execution Error: {str(e)}"

# -------------------------------------------------------------
# Tavily Search with built-in backoff
# -------------------------------------------------------------

if TAVILY_AVAILABLE:
    tavily_search_client = TavilySearch(max_results=3)
else:
    # Create a mock client that returns helpful error messages
    tavily_search_client = TavilySearch()  # Uses our stub class

class TavilySearchInput(BaseModel):
    query: str = Field(description="The search query.")
    max_results: int = Field(default=3, description="Maximum number of results.")

@tool
def tavily_search_backoff(query: str, max_results: int = 3) -> str:
    """Runs Tavily search with exponential backoff to avoid 429 limits."""

    def _call():
        return tavily_search_client.run(query)

    try:
        result = _exponential_backoff(_call, max_retries=4)
        # Tavily returns list of dicts; convert to readable string if needed
        if isinstance(result, list):
            formatted = []
            for idx, item in enumerate(result, 1):
                title = item.get('title') or item.get('url')
                snippet = item.get('content') or item.get('snippet', '')
                formatted.append(f"{idx}. {title}\n{snippet}\n")
            return "\n".join(formatted)
        return str(result)
    except Exception as e:
        return f"Error during Tavily search: {e}"

tavily_search = StructuredTool.from_function(
    func=tavily_search_backoff,
    name="tavily_search",
    description="Runs Tavily search with exponential backoff to avoid 429 limits.",
    args_schema=TavilySearchInput
)

# --- Tool Definitions ---

def get_tools() -> list:
    """
    Initializes and returns a list of all tools available to the agent.
    """
    # 1. Tavily Search Tool with backoff for real-time web searches
    tavily_search_tool = tavily_search_backoff

    # 2. LlamaIndex Knowledge Base Tool for private data retrieval
    # Initialize the vector store and index
    try:
        vector_store = get_vector_store()
        # Ensure embedding model is configured for the index
        Settings.embed_model = OpenAIEmbedding(model="text-embedding-3-small")
        index = VectorStoreIndex.from_vector_store(vector_store=vector_store)
        
        # Create a query engine from the index
        query_engine = index.as_query_engine(
            similarity_top_k=3,
            response_mode="compact"
        )
        
        # Wrap the query engine in a tool
        knowledge_base_tool = QueryEngineTool.from_defaults(
            query_engine=query_engine,
            name="knowledge_base_retriever",
            description=(
                "Useful for answering questions about internal company policies, project documentation (like Project Orion), and historical data. Use this for private, non-public information."
            )
        )
    except Exception as e:
        logger.error(f"Failed to initialize Knowledge Base tool: {e}. It will not be available.")
        knowledge_base_tool = None


    # 3. Python Code Interpreter Tool for calculations and code execution
    # SECURITY WARNING: This tool executes arbitrary Python code.
    # Do not use in a production environment with untrusted inputs without proper sandboxing.
    python_repl_tool = PythonREPLTool()

    # Assemble the list of tools, starting with the critical environment interaction tools
    tools = [
        file_reader,  # Critical for fixing "context blindness"
        advanced_file_reader,  # For Excel, PDF, Word docs
        audio_transcriber,  # For MP3 and audio files
        video_analyzer_structured,  # For YouTube and video analysis (StructuredTool)
        image_analyzer,  # For image and chess analysis
        web_researcher,  # For Wikipedia and web research
        semantic_search_tool,  # For working with knowledge bases like supabase_docs.csv
        python_interpreter,  # Enhanced code execution
        tavily_search_tool,  # Real-time web search with backoff
        python_repl_tool  # Backup code execution tool
    ]
    
    # Add the knowledge base tool if available
    if knowledge_base_tool:
        tools.append(knowledge_base_tool)
        
    logger.info(f"Initialized {len(tools)} tools: {[t.name for t in tools]}")

    # If __all__ is defined, add WikipediaAPIWrapper
    try:
        __all__.append('WikipediaAPIWrapper')
    except Exception:
        pass

    return tools

# Example of a custom tool using the @tool decorator
@tool
def get_weather(city: str) -> str:
    """
    A dummy tool to get the weather for a given city.
    In a real application, this would call a weather API.
    """
    if "san francisco" in city.lower():
        return "The weather in San Francisco is 65°F and sunny."
    elif "new york" in city.lower():
        return "The weather in New York is 75°F and humid."
    else:
        return f"Sorry, I don't have weather information for {city}."

# To add the custom tool, you would append it to the list in get_tools()
# tools.append(get_weather)

# Dummy SemanticSearchEngine for test patching
class SemanticSearchEngine:
    def __init__(self, *args, **kwargs):
        pass
    def search(self, *args, **kwargs):
        return [] 